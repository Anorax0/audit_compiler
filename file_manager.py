"""
Module handling excel file
"""

import sys
import warnings
from PyQt5.QtWidgets import QMessageBox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import colorama
from record_checker import RecordChecker

# initialize colorama
colorama.init()


def warning_box(title, message):
    """
    Displays warning box
    :param title: str
    :param message: str
    :return: None
    """

    warning = QMessageBox()
    warning.setWindowTitle(title)
    warning.setText(message)
    warning.exec_()


class FileManager:
    """
    Class manager for excel file
    """

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = Workbook()

    def load_file(self):
        """
        Load the file
        :return: workbook object
        """

        try:

            warnings.simplefilter("ignore")
            workbook = load_workbook(filename=self.file_path)
            workbook = workbook['Sheet1']
            warnings.simplefilter("default")

            return workbook

        except FileNotFoundError:
            warning_box('File not found!', 'File not found.')
            sys.exit(1)
        except PermissionError:
            warning_box('Permision Error!',
                        'File is running in different process. '
                        'Please close other processes to work with this file.')
            sys.exit(1)

    def check_col_jk(self):
        """
        Checks column J and K in workbook for incorrect records
        :return: boolean
        """

        workbook = load_workbook(self.file_path)
        ws = workbook['Sheet1']

        i = 0
        for row in range(2, self.get_max_row() + 1):
            if str(ws[f'J{row}'].value) == 'REVIEWED' or str(ws[f'J{row}'].value) == 'PENDING':
                pass
            else:
                pattern = PatternFill("solid", fgColor="FF0000")
                for cell in ws[row:row]:
                    cell.fill = pattern
                i += 1
        try:
            workbook.save(self.file_path)
        except PermissionError:
            warning_box('Permision Error!',
                        'File is running in different process. '
                        'Please close other processes to work with this file.')
            sys.exit(1)

        if i > 0:
            return True
        # else:
        #     return False

    def get_max_row(self):
        """
        Get int of last row in workbook
        :return: int
        """

        workbook = self.load_file()
        return workbook.max_row

    def save_workbook(self, records_list):
        """
        Saves the workbook
        :param records_list: list
        :return: None
        """

        warnings.simplefilter("ignore")
        try:
            ws = load_workbook(self.file_path)
            if 'Audit' in ws.sheetnames:
                to_remove = ws.get_sheet_by_name('Audit')
                ws.remove_sheet(to_remove)
                ws.create_sheet('Audit')
                audit_sheet = ws['Audit']
            else:
                ws.create_sheet('Audit')
                audit_sheet = ws['Audit']

            audit_sheet['A1'] = 'ASIN'
            audit_sheet['B1'] = 'Subject to Clp'
            audit_sheet['C1'] = 'Is kit'
            audit_sheet['D1'] = 'eu2008_labeling_risk'
            audit_sheet['E1'] = 'eu2008_labeling_hazard'
            audit_sheet['F1'] = 'SDS'
            audit_sheet['G1'] = 'TT'
            audit_sheet['H1'] = 'Reviewer Login'
            audit_sheet['I1'] = 'Review Date'
            audit_sheet['J1'] = 'State'
            audit_sheet['K1'] = 'Utc Reason'
            audit_sheet['L1'] = 'Is Dg Utc'
            audit_sheet['M1'] = 'CLP Exemption'

            for i, row in enumerate(records_list, start=2):
                checker = RecordChecker(ws['Sheet1'], row)
                audit_sheet[f'A{i}'] = checker.col_a
                audit_sheet[f'B{i}'] = checker.col_b
                audit_sheet[f'C{i}'] = checker.col_c
                audit_sheet[f'D{i}'] = checker.col_d
                audit_sheet[f'E{i}'] = checker.col_e
                audit_sheet[f'F{i}'] = checker.col_f
                audit_sheet[f'G{i}'] = checker.col_g
                audit_sheet[f'H{i}'] = checker.col_h
                audit_sheet[f'I{i}'] = checker.col_i
                audit_sheet[f'J{i}'] = checker.col_j
                audit_sheet[f'K{i}'] = checker.col_k
                audit_sheet[f'L{i}'] = checker.col_l
                audit_sheet[f'M{i}'] = checker.col_m
            ws.save(self.file_path)
        except PermissionError:
            warning_box('Permision Error!',
                        'File is running in different process. '
                        'Please close other processes to work with this file.')
            sys.exit(1)
        except BaseException as error:
            print(error)
            warning_box('Error!',
                        'Something gone wrong during saving records to new worksheet.')
            sys.exit(1)
        finally:
            warnings.simplefilter("default")


if __name__ == '__main__':
    file = 'test.xlsx'
    wb = FileManager(file)
    wkb = wb.load_file()
    records = [168, 147, 144, 107, 135, 142, 248]
    wb.save_workbook(records)
